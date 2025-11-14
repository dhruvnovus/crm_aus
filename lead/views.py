from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q, Count, Sum, F, Value, CharField
from django.db.models.functions import Concat
from django.http import Http404
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiParameter, OpenApiExample
from drf_spectacular.types import OpenApiTypes

from .models import Lead, LeadHistory, RegistrationGroup, LeadTag, SponsorshipType
from .serializers import (
    LeadListSerializer,
    LeadDetailSerializer,
    LeadCreateUpdateSerializer,
    LeadStatsSerializer,
    LeadBulkImportSerializer,
    LeadHistorySerializer,
    RegistrationGroupSerializer,
    LeadTagSerializer,
    SponsorshipTypeSerializer
)


class LeadPagination(PageNumberPagination):
    """
    Custom pagination for lead list
    """
    page_size = 100
    page_size_query_param = 'page_size'
    max_page_size = 500


@extend_schema_view(
    list=extend_schema(
        summary="List all leads",
        description="Get a paginated list of all leads with optional filtering",
        tags=["Leads"],
    ),
    create=extend_schema(
        summary="Create lead",
        description="Create a new lead",
        tags=["Leads"],
    ),
    retrieve=extend_schema(
        summary="Get lead details",
        description="Retrieve detailed information about a specific lead",
        tags=["Leads"],
    ),
    update=extend_schema(
        summary="Update lead (full)",
        description="Update all fields of a lead",
        tags=["Leads"],
    ),
    partial_update=extend_schema(
        summary="Update lead (partial)",
        description="Update specific fields of a lead",
        tags=["Leads"],
    ),
    destroy=extend_schema(
        summary="Delete lead",
        description="Delete a lead from the system",
        tags=["Leads"],
    ),
)
class LeadViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Lead CRUD operations
    Provides: list, create, retrieve, update, partial_update, destroy
    """
    queryset = Lead.objects.select_related('assigned_sales_staff').all()
    pagination_class = LeadPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = [
        'status', 'lead_type', 'intensity', 'assigned_sales_staff', 'event',
        'sponsorship_type', 'registration_groups', 'tags', 'lead_name', 'lead_pipeline', 'lead_stage'
    ]
    search_fields = [
        'first_name', 'last_name', 'company_name', 'email_address', 'contact_number',
        'tags__name', 'sponsorship_type__name', 'registration_groups__name',
        'lead_name', 'lead_pipeline', 'lead_stage'
    ]
    ordering_fields = ['date_received', 'created_at', 'updated_at', 'first_name', 'last_name', 'full_name', 'full_name_ordering', 'company_name', 'opportunity_price']
    ordering = ['-date_received']

    @staticmethod
    def _first_error(errors):
        try:
            first = next(iter(errors.values()))
            # errors may be list/tuple/dict
            while isinstance(first, (list, tuple)) and first:
                first = first[0]
            if isinstance(first, dict):
                first = next(iter(first.values()))
            return str(first)
        except Exception:
            return "Invalid data"
    
    def get_serializer_class(self):
        """
        Return appropriate serializer class based on action
        """
        if self.action == 'list':
            return LeadListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return LeadCreateUpdateSerializer
        else:
            return LeadDetailSerializer
    
    def perform_create(self, serializer):
        """Create lead and send notification if assigned"""
        lead = serializer.save()
        
        # Create notification if lead is assigned during creation
        if lead.assigned_sales_staff:
            try:
                from notifications.signals import create_lead_assignment_notification
                create_lead_assignment_notification(lead, lead.assigned_sales_staff)
            except Exception as e:
                # Log error but don't fail lead creation
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Failed to create notification for lead {lead.id}: {str(e)}", exc_info=True)
    
    def perform_update(self, serializer):
        """Update lead and send notification if assignment changed"""
        old_assigned = serializer.instance.assigned_sales_staff
        lead = serializer.save()
        
        # Create notification if assignment changed
        if lead.assigned_sales_staff and old_assigned != lead.assigned_sales_staff:
            try:
                from notifications.signals import create_lead_assignment_notification
                create_lead_assignment_notification(lead, lead.assigned_sales_staff)
            except Exception as e:
                # Log error but don't fail lead update
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Failed to create notification for lead {lead.id}: {str(e)}", exc_info=True)
    
    def get_queryset(self):
        """
        Optionally restricts the returned leads by filtering against
        query parameters in the URL.
        """
        # Filter out deleted records by default and optimize with select_related
        # Annotate full_name_ordering for database-level ordering (different name to avoid @property conflict)
        queryset = Lead.objects.select_related('assigned_sales_staff').filter(is_deleted=False).annotate(
            full_name_ordering=Concat(
                F('first_name'),
                Value(' '),
                F('last_name'),
                output_field=CharField()
            )
        )
        
        # Filter by status category
        status_category = self.request.query_params.get('status_category', None)
        if status_category:
            if status_category == 'active':
                queryset = queryset.exclude(status__in=['lost', 'withdrawn'])
            elif status_category == 'inactive':
                queryset = queryset.filter(status__in=['lost', 'withdrawn'])
        
        return queryset
    
    def filter_queryset(self, queryset):
        """
        Override to handle full_name ordering by mapping it to full_name_ordering annotation
        """
        # Get ordering parameter
        ordering_param = self.request.query_params.get('ordering', '')
        
        # If ordering includes full_name, handle it manually
        # because full_name is a @property, not a database field
        if ordering_param and 'full_name' in ordering_param and 'full_name_ordering' not in ordering_param:
            # Temporarily remove full_name from ordering_fields to avoid OrderingFilter error
            original_ordering_fields = self.ordering_fields
            self.ordering_fields = [f for f in self.ordering_fields if f != 'full_name']
            
            # Call super() to apply other filters (without full_name ordering)
            queryset = super().filter_queryset(queryset)
            
            # Restore ordering_fields
            self.ordering_fields = original_ordering_fields
            
            # Now apply full_name ordering manually using full_name_ordering annotation
            # Replace full_name with full_name_ordering in the ordering string
            modified_ordering = ordering_param.replace('full_name', 'full_name_ordering')
            # Split by comma to handle multiple ordering fields
            order_fields = [field.strip() for field in modified_ordering.split(',')]
            queryset = queryset.order_by(*order_fields)
        else:
            # No full_name in ordering, use normal flow
            queryset = super().filter_queryset(queryset)
        
        return queryset
    
    def list(self, request, *args, **kwargs):
        """
        List leads with optional filtering
        """
        queryset = self.filter_queryset(self.get_queryset())
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    def create(self, request, *args, **kwargs):
        """
        Create a new lead
        """
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            return Response({"status": False, "error": self._first_error(serializer.errors)}, status=status.HTTP_400_BAD_REQUEST)
        try:
            # Use perform_create to ensure notification is created
            self.perform_create(serializer)
            lead = serializer.instance
        except Exception as exc:
            return Response({"status": False, "error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        # Return detailed lead data
        detail_serializer = LeadDetailSerializer(lead)
        return Response({"success": True, "message": "Lead created successfully", "data": detail_serializer.data}, status=status.HTTP_201_CREATED)
    
    
    def update(self, request, *args, **kwargs):
        """
        Update a lead (full update)
        """
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if not serializer.is_valid():
            return Response({"status": False, "error": self._first_error(serializer.errors)}, status=status.HTTP_400_BAD_REQUEST)
        try:
            # Use perform_update to ensure notification is created if assignment changed
            self.perform_update(serializer)
            lead = serializer.instance
        except Exception as exc:
            return Response({"status": False, "error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        # Return detailed lead data
        detail_serializer = LeadDetailSerializer(lead)
        return Response({"success": True, "message": "Lead updated successfully", "data": detail_serializer.data})
    
    def partial_update(self, request, *args, **kwargs):
        """
        Partially update a lead
        """
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        """
        Soft delete a lead (sets is_deleted=True instead of actually deleting)
        """
        try:
            instance = self.get_object()
            if hasattr(instance, 'is_deleted'):
                # Use queryset update to avoid model save hooks
                Lead.objects.filter(pk=instance.pk).update(is_deleted=True)
            else:
                # Fallback if column not present yet
                self.perform_destroy(instance)
            return Response({"success": True, "message": "Lead deleted successfully"}, status=status.HTTP_200_OK)
        except Exception as exc:
            return Response({"success": False, "error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
    
    
    @extend_schema(
        summary="Get lead statistics",
        description="Get comprehensive statistics about leads (excludes deleted records)",
        tags=["Leads"],
    )
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Get lead statistics (excludes deleted records)
        """
        queryset = Lead.objects.filter(is_deleted=False)
        
        stats = {
            'total_leads': queryset.count(),
            'new_leads': queryset.filter(status='new').count(),
            'info_pack_leads': queryset.filter(status='info_pack').count(),
            'attempted_contact_leads': queryset.filter(status='attempted_contact').count(),
            'contacted_leads': queryset.filter(status='contacted').count(),
            'contract_invoice_sent_leads': queryset.filter(status='contract_invoice_sent').count(),
            'contract_signed_paid_leads': queryset.filter(status='contract_signed_paid').count(),
            'withdrawn_leads': queryset.filter(status='withdrawn').count(),
            'lost_leads': queryset.filter(status='lost').count(),
            'converted_leads': queryset.filter(status='converted').count(),
            'future_leads': queryset.filter(status='future').count(),
            'total_opportunity_value': queryset.aggregate(
                total=Sum('opportunity_price')
            )['total'] or 0,
            'exhibitor_count': queryset.filter(lead_type='exhibitor').count(),
            'sponsor_count': queryset.filter(lead_type='sponsor').count(),
            'visitor_count': queryset.filter(lead_type='visitor').count(),
        }
        
        serializer = LeadStatsSerializer(stats)
        return Response(serializer.data)
    
    @extend_schema(
        summary="Get leads by status",
        description="Get leads filtered by specific status",
        tags=["Leads"],
    )
    @action(detail=False, methods=['get'])
    def by_status(self, request):
        """
        Get leads by specific status
        """
        status_param = request.query_params.get('status', 'new')
        queryset = self.get_queryset().filter(status=status_param)
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @extend_schema(
        summary="Get new leads",
        description="Get all new leads",
        tags=["Leads"],
    )
    @action(detail=False, methods=['get'])
    def new_leads(self, request):
        """
        Get all new leads
        """
        queryset = self.get_queryset().filter(status='new')
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @extend_schema(
        summary="Get lost leads",
        description="Get all lost leads",
        tags=["Leads"],
    )
    @action(detail=False, methods=['get'])
    def lost_leads(self, request):
        """
        Get all lost leads
        """
        queryset = self.get_queryset().filter(status='lost')
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @extend_schema(
        summary="Get converted leads",
        description="Get all converted leads",
        tags=["Leads"],
    )
    @action(detail=False, methods=['get'])
    def converted_leads(self, request):
        """
        Get all converted leads
        """
        queryset = self.get_queryset().filter(status='converted')
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @extend_schema(
        summary="Get future leads",
        description="Get all future leads",
        tags=["Leads"],
    )
    @action(detail=False, methods=['get'])
    def future_leads(self, request):
        """
        Get all future leads
        """
        queryset = self.get_queryset().filter(status='future')
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @extend_schema(
        summary="Update lead status",
        description="Update the status of a specific lead",
        tags=["Leads"],
    )
    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        """
        Update lead status
        """
        lead = self.get_object()
        new_status = request.data.get('status')
        
        if not new_status:
            return Response(
                {"error": "Status is required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if new_status not in [choice[0] for choice in Lead.STATUS_CHOICES]:
            return Response(
                {"error": "Invalid status."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        lead.status = new_status
        lead.save()
        
        serializer = LeadDetailSerializer(lead)
        return Response(serializer.data)
    
    @extend_schema(
        summary="Assign lead to sales staff",
        description="Assign a lead to a specific sales staff member. Requires employee_id in request body.",
        tags=["Leads"],
        request={
            'application/json': {
                'type': 'object',
                'properties': {
                    'employee_id': {
                        'type': 'integer',
                        'description': 'ID of the employee to assign the lead to'
                    }
                },
                'required': ['employee_id']
            }
        }
    )
    @action(detail=True, methods=['post'])
    def assign_sales_staff(self, request, pk=None):
        """
        Assign lead to sales staff
        Requires only employee_id in request body
        """
        try:
            lead = self.get_object()
            employee_id = request.data.get('employee_id')
            
            if not employee_id:
                return Response(
                    {"status": False, "error": "employee_id is required."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            try:
                from employee.models import Employee
                employee = Employee.objects.get(id=employee_id, is_deleted=False)
            except Employee.DoesNotExist:
                return Response(
                    {"status": False, "error": "Employee not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            old_assigned = lead.assigned_sales_staff
            lead.assigned_sales_staff = employee
            lead.save()
            
            # Create notification if assignment changed and is not empty
            if employee and old_assigned != employee:
                try:
                    from notifications.signals import create_lead_assignment_notification
                    create_lead_assignment_notification(lead, employee)
                except Exception as e:
                    # Log error but don't fail the assignment
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f"Failed to create notification for lead assignment: {str(e)}")
            
            serializer = LeadDetailSerializer(lead)
            return Response({"success": True, "message": "Lead assigned successfully", "data": serializer.data})
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error in assign_sales_staff: {str(e)}", exc_info=True)
            return Response(
                {"status": False, "error": f"An error occurred while assigning lead: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @extend_schema(
        summary="Bulk import leads",
        description="Import multiple leads at once",
        tags=["Leads"],
        examples=[
            OpenApiExample(
                "Bulk Import Example",
                summary="Import multiple leads",
                description="Example of importing multiple leads",
                value={
                    "leads_data": [
                        {
                            "title": "mr",
                            "first_name": "John",
                            "last_name": "Doe",
                            "company_name": "Example Company",
                            "contact_number": "+61412345678",
                            "email_address": "john.doe@example.com",
                            "lead_type": "exhibitor",
                            "status": "new",
                            "event": "Aged & Disability Expo Newcastle"
                        },
                        {
                            "title": "mrs",
                            "first_name": "Jane",
                            "last_name": "Smith",
                            "company_name": "Another Company",
                            "contact_number": "+61412345679",
                            "email_address": "jane.smith@example.com",
                            "lead_type": "sponsor",
                            "status": "new",
                            "event": "Aged & Disability Expo Sydney"
                        }
                    ]
                }
            )
        ]
    )
    @action(detail=False, methods=['post'])
    def bulk_import(self, request):
        """
        Bulk import leads
        """
        serializer = LeadBulkImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        leads_data = serializer.validated_data['leads_data']
        created_leads = []
        errors = []
        
        for lead_data in leads_data:
            try:
                lead_serializer = LeadCreateUpdateSerializer(data=lead_data)
                if lead_serializer.is_valid():
                    lead = lead_serializer.save()
                    created_leads.append(lead)
                else:
                    errors.append({
                        'data': lead_data,
                        'errors': lead_serializer.errors
                    })
            except Exception as e:
                errors.append({
                    'data': lead_data,
                    'errors': str(e)
                })
        
        response_data = {
            'created_count': len(created_leads),
            'error_count': len(errors),
            'created_leads': LeadDetailSerializer(created_leads, many=True).data,
            'errors': errors
        }
        
        return Response(response_data, status=status.HTTP_201_CREATED)
    
    @extend_schema(
        summary="Export leads",
        description="Export leads to CSV format",
        tags=["Leads"],
    )
    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        Export leads to CSV (excludes deleted records by default)
        """
        import csv
        from django.http import HttpResponse
        
        queryset = self.filter_queryset(self.get_queryset())
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="leads_export.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'ID', 'Title', 'First Name', 'Last Name', 'Company Name',
            'Contact Number', 'Email Address', 'Custom Email Addresses',
            'Address', 'Event', 'Lead Type', 'Booth Size', 'Sponsorship Type',
            'Registration Groups', 'Status', 'Intensity', 'Opportunity Price',
            'Tags', 'How Did You Hear', 'Reason for Enquiry',
            'Assigned Sales Staff', 'Lead Name', 'Lead Pipeline', 'Lead Stage',
            'Date Received', 'Created At', 'Updated At'
        ])
        
        # Write data
        for lead in queryset:
            sponsorship_names = ", ".join([s.name for s in lead.sponsorship_type.all()])
            registration_names = ", ".join([r.name for r in lead.registration_groups.all()])
            tag_names = ", ".join([t.name for t in lead.tags.all()])
            writer.writerow([
                lead.id, lead.get_title_display(), lead.first_name, lead.last_name,
                lead.company_name, lead.contact_number, lead.email_address,
                lead.custom_email_addresses, lead.address, lead.event,
                lead.get_lead_type_display(), lead.booth_size, sponsorship_names,
                registration_names, lead.get_status_display(),
                lead.get_intensity_display(), lead.opportunity_price, tag_names,
                lead.how_did_you_hear, lead.reason_for_enquiry, lead.assigned_sales_staff.full_name if lead.assigned_sales_staff else '',
                lead.lead_name, lead.lead_pipeline, lead.lead_stage,
                lead.date_received, lead.created_at, lead.updated_at
            ])
        
        return response
    
    @extend_schema(
        summary="Import leads from CSV/Excel",
        description="Import leads from a CSV or Excel file. Supports both .csv and .xlsx formats. Use multipart/form-data with 'file' field.",
        tags=["Leads"],
        request={
            'multipart/form-data': {
                'type': 'object',
                'properties': {
                    'file': {
                        'type': 'string',
                        'format': 'binary',
                        'description': 'CSV or Excel file (.csv or .xlsx) containing lead data'
                    }
                },
                'required': ['file']
            }
        },
        responses={
            200: {
                'description': 'Import results',
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'message': {'type': 'string'},
                    'created_count': {'type': 'integer'},
                    'error_count': {'type': 'integer'},
                    'errors': {'type': 'array'}
                }
            }
        }
    )
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def import_leads(self, request):
        """
        Import leads from CSV or Excel file
        """
        import csv
        import io
        from openpyxl import load_workbook
        
        if 'file' not in request.FILES:
            return Response(
                {"success": False, "error": "No file provided. Please upload a CSV or Excel file."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        uploaded_file = request.FILES['file']
        file_name = uploaded_file.name.lower()
        
        # Validate file extension
        if not (file_name.endswith('.csv') or file_name.endswith('.xlsx') or file_name.endswith('.xls')):
            return Response(
                {"success": False, "error": "Invalid file format. Please upload a CSV (.csv) or Excel (.xlsx, .xls) file."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        rows = []
        
        try:
            # Read CSV file
            if file_name.endswith('.csv'):
                # Decode file content
                file_content = uploaded_file.read().decode('utf-8-sig')  # Handle BOM
                csv_reader = csv.DictReader(io.StringIO(file_content))
                all_csv_rows = list(csv_reader)
                
                # Filter out empty rows - only keep rows with at least one required field
                # Required fields: Name (or first_name/last_name), Email, Mobile No
                required_fields = ['name', 'first_name', 'last_name', 'email', 'email_address', 
                                 'mobile', 'mobile_no', 'mobile no', 'contact_number', 'contact number',
                                 'phone', 'phone_number', 'phone number']
                
                rows = []
                for row in all_csv_rows:
                    has_required_data = False
                    for key, value in row.items():
                        key_lower = str(key).lower().strip()
                        value_str = str(value).strip() if value else ''
                        # Check if this is a required field and has meaningful data
                        if any(req_field in key_lower for req_field in required_fields):
                            if value_str and value_str.lower() not in ['', 'none', 'null', 'n/a', 'na']:
                                has_required_data = True
                                break
                    if has_required_data:
                        rows.append(row)
            
            # Read Excel file
            elif file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                # Read file content into memory
                file_content = uploaded_file.read()
                workbook = load_workbook(filename=io.BytesIO(file_content), data_only=True)
                sheet = workbook.active
                
                # Get headers from first row
                headers = []
                for cell in sheet[1]:
                    if cell.value:
                        headers.append(str(cell.value).strip())
                    else:
                        headers.append('')
                
                # Read data rows
                for row in sheet.iter_rows(min_row=2, values_only=False):
                    row_data = {}
                    for idx, cell in enumerate(row):
                        if idx < len(headers) and headers[idx]:
                            value = cell.value
                            # Convert to string, handling None
                            if value is not None:
                                row_data[headers[idx]] = str(value).strip()
                            else:
                                row_data[headers[idx]] = ''
                    
                    # Only add row if it has at least one required field with non-empty value
                    # Required fields: Name (or first_name/last_name), Email, Mobile No
                    required_fields = ['name', 'first_name', 'last_name', 'email', 'email_address', 
                                     'mobile', 'mobile_no', 'mobile no', 'contact_number', 'contact number',
                                     'phone', 'phone_number', 'phone number']
                    
                    # Check if row has at least one required field with actual data (not empty, not "N/A", not "None")
                    has_required_data = False
                    for key, value in row_data.items():
                        key_lower = str(key).lower().strip()
                        value_str = str(value).strip() if value else ''
                        # Check if this is a required field and has meaningful data
                        if any(req_field in key_lower for req_field in required_fields):
                            if value_str and value_str.lower() not in ['', 'none', 'null', 'n/a', 'na']:
                                has_required_data = True
                                break
                    
                    if has_required_data:
                        rows.append(row_data)
            
            if not rows:
                return Response(
                    {"success": False, "error": "File is empty or contains no data."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Process rows and create leads
            created_leads = []
            errors = []
            
            for row_num, row_data in enumerate(rows, start=2):  # Start at 2 (1 is header)
                try:
                    # Map CSV/Excel columns to Lead model fields
                    try:
                        lead_data = self._map_row_to_lead_data(row_data)
                    except Exception as map_error:
                        errors.append({
                            'row': row_num,
                            'data': row_data,
                            'errors': f"Error mapping row data: {str(map_error)}"
                        })
                        continue
                    
                    # Validate and create lead
                    lead_serializer = LeadCreateUpdateSerializer(data=lead_data)
                    if lead_serializer.is_valid():
                        try:
                            lead = lead_serializer.save()
                            created_leads.append(lead)
                        except Exception as save_error:
                            errors.append({
                                'row': row_num,
                                'data': row_data,
                                'errors': f"Error saving lead: {str(save_error)}"
                            })
                    else:
                        errors.append({
                            'row': row_num,
                            'data': row_data,
                            'errors': lead_serializer.errors
                        })
                except Exception as e:
                    errors.append({
                        'row': row_num,
                        'data': row_data,
                        'errors': f"Unexpected error: {str(e)}"
                    })
            
            response_data = {
                'success': True,
                'message': f'Import completed. {len(created_leads)} leads created, {len(errors)} errors.',
                'created_count': len(created_leads),
                'error_count': len(errors),
                'errors': errors[:50]  # Limit to first 50 errors to avoid huge response
            }
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error importing leads: {str(e)}", exc_info=True)
            return Response(
                {
                    "success": False, 
                    "error": f"Error processing file: {str(e)}"
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _map_row_to_lead_data(self, row_data):
        """
        Map CSV/Excel row data to Lead model fields
        Handles column name variations and relationships
        """
        # Normalize keys (case-insensitive, strip whitespace, replace spaces with underscores)
        normalized_row = {}
        for k, v in row_data.items():
            try:
                # Normalize key: lowercase, strip, replace spaces/hyphens with underscores
                normalized_key = str(k).strip().lower().replace(' ', '_').replace('-', '_')
                # Handle value - convert to string, handle None, "None", "null", etc.
                if v is None:
                    normalized_row[normalized_key] = ''
                else:
                    v_str = str(v).strip()
                    # Treat "None", "null", "NULL" as empty
                    if v_str.lower() in ['none', 'null', 'nan', '']:
                        normalized_row[normalized_key] = ''
                    else:
                        normalized_row[normalized_key] = v_str
            except Exception as e:
                # Skip problematic keys
                continue
        
        # Also create aliases for common variations (without underscores)
        aliases = {}
        for key, value in normalized_row.items():
            # Create alias without underscores
            alias_key = key.replace('_', '')
            if alias_key not in normalized_row:
                aliases[alias_key] = value
        normalized_row.update(aliases)
        
        lead_data = {}
        
        # Default title
        lead_data['title'] = 'mr'  # Default
        
        # Handle "Name" field - split into first_name and last_name
        name_field = (
            normalized_row.get('name', '') or 
            normalized_row.get('full_name', '') or
            normalized_row.get('fullname', '') or
            normalized_row.get('full name', '')
        ).strip()
        
        # Also check for separate first_name and last_name fields
        first_name_raw = str(
            normalized_row.get('first_name', '') or 
            normalized_row.get('firstname', '') or
            normalized_row.get('first name', '') or
            normalized_row.get('fname', '') or ''
        ).strip()
        
        last_name_raw = str(
            normalized_row.get('last_name', '') or 
            normalized_row.get('lastname', '') or
            normalized_row.get('last name', '') or
            normalized_row.get('lname', '') or
            normalized_row.get('surname', '') or ''
        ).strip()
        
        # If "Name" field exists, split it into first and last name
        if name_field and not first_name_raw and not last_name_raw:
            # Split name by space - first word is first name, rest is last name
            name_parts = name_field.split()
            if len(name_parts) >= 1:
                first_name_raw = name_parts[0].strip()
                # Handle "N/A" in name
                if ' N/A' in first_name_raw or ' n/a' in first_name_raw:
                    first_name_raw = first_name_raw.split(' N/A')[0].split(' n/a')[0].strip()
            if len(name_parts) >= 2:
                last_name_raw = ' '.join(name_parts[1:]).strip()
            else:
                last_name_raw = ''
        
        # Handle cases where first_name might contain "N/A" or be split
        if first_name_raw and (' N/A' in first_name_raw or ' n/a' in first_name_raw):
            first_name_raw = first_name_raw.split(' N/A')[0].split(' n/a')[0].strip()
        
        lead_data['first_name'] = first_name_raw if first_name_raw else 'N/A'
        
        # Handle "None" string, empty, or None value for last name
        if not last_name_raw or last_name_raw.lower() in ['none', 'null', 'n/a', 'na']:
            lead_data['last_name'] = 'N/A'
        else:
            lead_data['last_name'] = last_name_raw
        
        # Company name (required) - default to "N/A" if not provided
        company_name_raw = str(
            normalized_row.get('company_name', '') or 
            normalized_row.get('companyname', '') or
            normalized_row.get('company name', '') or
            normalized_row.get('company', '') or ''
        ).strip()
        lead_data['company_name'] = company_name_raw if company_name_raw else 'N/A'
        
        # Contact number (required) - map from "Mobile No" or "Mobile No"
        contact_number_raw = str(
            normalized_row.get('mobile_no', '') or 
            normalized_row.get('mobile no', '') or
            normalized_row.get('mobileno', '') or
            normalized_row.get('contact_number', '') or 
            normalized_row.get('contactnumber', '') or
            normalized_row.get('contact number', '') or
            normalized_row.get('phone', '') or
            normalized_row.get('mobile', '') or
            normalized_row.get('phone_number', '') or
            normalized_row.get('phone number', '') or
            normalized_row.get('mobile_number', '') or
            normalized_row.get('mobile number', '') or
            normalized_row.get('contact', '') or ''
        ).strip()
        # Contact number must match regex: ^\+?1?\d{9,15}$
        if contact_number_raw:
            lead_data['contact_number'] = contact_number_raw
        else:
            # Use a valid default that passes the regex validator
            lead_data['contact_number'] = '1234567890'
        
        # Email address (required) - map from "Email"
        email_address_raw = str(
            normalized_row.get('email', '') or
            normalized_row.get('email_address', '') or 
            normalized_row.get('emailaddress', '') or
            normalized_row.get('email address', '') or
            normalized_row.get('e_mail', '') or
            normalized_row.get('e mail', '') or ''
        ).strip().lower()
        lead_data['email_address'] = email_address_raw if email_address_raw else 'noemail@example.com'
        
        # Optional fields
        custom_emails = (
            normalized_row.get('custom_email_addresses', '') or 
            normalized_row.get('customemailaddresses', '') or
            normalized_row.get('custom email addresses', '') or
            normalized_row.get('custom_emails', '') or
            normalized_row.get('custom emails', '')
        )
        if custom_emails:
            lead_data['custom_email_addresses'] = custom_emails
        
        if normalized_row.get('address', ''):
            lead_data['address'] = normalized_row.get('address', '')
        
        if normalized_row.get('event', ''):
            lead_data['event'] = normalized_row.get('event', '')
        
        # Lead type
        lead_type = (
            normalized_row.get('lead_type', '') or 
            normalized_row.get('leadtype', '') or
            normalized_row.get('lead type', '') or
            normalized_row.get('type', '')
        ).lower()
        if lead_type in ['exhibitor', 'sponsor', 'visitor']:
            lead_data['lead_type'] = lead_type
        else:
            lead_data['lead_type'] = 'exhibitor'  # Default
        
        booth_size = (
            normalized_row.get('booth_size', '') or 
            normalized_row.get('boothsize', '') or
            normalized_row.get('booth size', '')
        )
        if booth_size:
            lead_data['booth_size'] = booth_size
        
        # Status
        status_value = (
            normalized_row.get('status', '') or 
            normalized_row.get('lead_status', '') or
            normalized_row.get('leadstatus', '') or
            normalized_row.get('lead status', '') or
            normalized_row.get('lead_status_display', '') or
            normalized_row.get('leadstatusdisplay', '') or
            normalized_row.get('lead status display', '')
        ).strip()
        
        if status_value:
            status_lower = status_value.lower()
            # First, check if it's a valid status key (e.g., 'new', 'contacted')
            valid_status_keys = [choice[0] for choice in Lead.STATUS_CHOICES]
            if status_lower in valid_status_keys:
                lead_data['status'] = status_lower
            else:
                # Try to match display names (e.g., 'New', 'Contacted', 'Info Pack')
                status_mapping = {
                    'new': 'new',
                    'attendee': 'attendee',
                    'job leads': 'job_leads',
                    'job_leads': 'job_leads',
                    'info pack': 'info_pack',
                    'info_pack': 'info_pack',
                    'attempted contact': 'attempted_contact',
                    'attempted_contact': 'attempted_contact',
                    'contacted': 'contacted',
                    'contract signed': 'contract_signed',
                    'contract_signed': 'contract_signed',
                    'contract & invoice sent': 'contract_invoice_sent',
                    'contract_invoice_sent': 'contract_invoice_sent',
                    'contract and invoice sent': 'contract_invoice_sent',
                    'contract signed & paid': 'contract_signed_paid',
                    'contract_signed_paid': 'contract_signed_paid',
                    'contract signed and paid': 'contract_signed_paid',
                    'withdrawn': 'withdrawn',
                    'lost': 'lost',
                    'converted': 'converted',
                    'future': 'future',
                }
                # Try exact match first
                if status_lower in status_mapping:
                    lead_data['status'] = status_mapping[status_lower]
                else:
                    # Try partial match (e.g., "Info Pack" -> "info pack")
                    matched = False
                    for display_name, status_key in status_mapping.items():
                        if display_name in status_lower or status_lower in display_name:
                            lead_data['status'] = status_key
                            matched = True
                            break
                    if not matched:
                        # Default to 'new' if no match found
                        lead_data['status'] = 'new'
        else:
            lead_data['status'] = 'new'  # Default
        
        # Intensity
        intensity_value = (
            normalized_row.get('intensity', '') or 
            normalized_row.get('lead_intensity', '') or
            normalized_row.get('leadintensity', '') or
            normalized_row.get('lead intensity', '')
        ).lower()
        if intensity_value in ['cold', 'warm', 'hot', 'sql']:
            lead_data['intensity'] = intensity_value
        else:
            lead_data['intensity'] = 'cold'  # Default
        
        if normalized_row.get('opportunity_price', ''):
            try:
                lead_data['opportunity_price'] = float(normalized_row.get('opportunity_price', '').replace(',', ''))
            except (ValueError, AttributeError):
                pass
        
        how_did_you_hear = (
            normalized_row.get('how_did_you_hear', '') or 
            normalized_row.get('howdidyouhear', '') or
            normalized_row.get('how did you hear', '') or
            normalized_row.get('how_did_you_hear_about_us', '') or
            normalized_row.get('howdidyouhearaboutus', '') or
            normalized_row.get('how did you hear about us', '')
        )
        if how_did_you_hear:
            lead_data['how_did_you_hear'] = how_did_you_hear
        
        reason_for_enquiry = (
            normalized_row.get('reason_for_enquiry', '') or 
            normalized_row.get('reasonforenquiry', '') or
            normalized_row.get('reason for enquiry', '') or
            normalized_row.get('reason', '')
        )
        if reason_for_enquiry:
            lead_data['reason_for_enquiry'] = reason_for_enquiry
        
        lead_name = (
            normalized_row.get('lead_name', '') or 
            normalized_row.get('leadname', '') or
            normalized_row.get('lead name', '')
        )
        if lead_name:
            lead_data['lead_name'] = lead_name
        
        lead_pipeline = (
            normalized_row.get('lead_pipeline', '') or 
            normalized_row.get('leadpipeline', '') or
            normalized_row.get('lead pipeline', '')
        )
        if lead_pipeline:
            lead_data['lead_pipeline'] = lead_pipeline
        
        lead_stage = (
            normalized_row.get('lead_stage', '') or 
            normalized_row.get('leadstage', '') or
            normalized_row.get('lead stage', '')
        )
        if lead_stage:
            lead_data['lead_stage'] = lead_stage
        
        # Handle ManyToMany relationships - create or get objects by name
        # Sponsorship Types (comma-separated)
        sponsorship_types = (
            normalized_row.get('sponsorship_type', '') or 
            normalized_row.get('sponsorshiptype', '') or
            normalized_row.get('sponsorship type', '') or
            normalized_row.get('sponsorship_types', '') or
            normalized_row.get('sponsorshiptypes', '') or
            normalized_row.get('sponsorship types', '')
        )
        if sponsorship_types:
            sponsorship_type_names = [s.strip() for s in sponsorship_types.split(',') if s.strip()]
            sponsorship_type_ids = []
            for name in sponsorship_type_names:
                sponsorship_type, created = SponsorshipType.objects.get_or_create(
                    name=name,
                    defaults={'is_deleted': False}
                )
                sponsorship_type_ids.append(sponsorship_type.id)
            if sponsorship_type_ids:
                lead_data['sponsorship_type'] = sponsorship_type_ids
        
        # Registration Groups (comma-separated)
        registration_groups = (
            normalized_row.get('registration_groups', '') or 
            normalized_row.get('registrationgroups', '') or
            normalized_row.get('registration groups', '') or
            normalized_row.get('registration_group', '') or
            normalized_row.get('registrationgroup', '') or
            normalized_row.get('registration group', '')
        )
        if registration_groups:
            registration_group_names = [r.strip() for r in registration_groups.split(',') if r.strip()]
            registration_group_ids = []
            for name in registration_group_names:
                registration_group, created = RegistrationGroup.objects.get_or_create(
                    name=name,
                    defaults={'is_deleted': False}
                )
                registration_group_ids.append(registration_group.id)
            if registration_group_ids:
                lead_data['registration_groups'] = registration_group_ids
        
        # Tags (comma-separated)
        tags = (
            normalized_row.get('tags', '') or 
            normalized_row.get('tag', '')
        )
        if tags:
            tag_names = [t.strip() for t in tags.split(',') if t.strip()]
            tag_ids = []
            for name in tag_names:
                tag, created = LeadTag.objects.get_or_create(
                    name=name,
                    defaults={'is_deleted': False}
                )
                tag_ids.append(tag.id)
            if tag_ids:
                lead_data['tags'] = tag_ids
        
        # Assigned Sales Staff (by employee ID or email - prefer ID)
        assigned_staff = (
            normalized_row.get('assigned_sales_staff', '') or 
            normalized_row.get('assignedsalesstaff', '') or
            normalized_row.get('assigned sales staff', '') or
            normalized_row.get('assigned_staff', '') or
            normalized_row.get('assignedstaff', '') or
            normalized_row.get('assigned staff', '') or
            normalized_row.get('employee_id', '') or
            normalized_row.get('employeeid', '') or
            normalized_row.get('employee id', '') or
            normalized_row.get('sales_staff', '') or
            normalized_row.get('salesstaff', '') or
            normalized_row.get('sales staff', '')
        )
        if assigned_staff:
            try:
                # Try as ID first
                employee_id = int(assigned_staff)
                from employee.models import Employee
                employee = Employee.objects.filter(id=employee_id, is_deleted=False).first()
                if employee:
                    lead_data['employee_id'] = employee_id
            except (ValueError, TypeError):
                # Try as email
                try:
                    from employee.models import Employee
                    employee = Employee.objects.filter(email=assigned_staff, is_deleted=False).first()
                    if employee:
                        lead_data['employee_id'] = employee.id
                except:
                    pass
        
        return lead_data

    # History endpoints
    @extend_schema(
        summary="List lead history entries",
        tags=["Leads"],
    )
    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        lead = self.get_object()
        queryset = LeadHistory.objects.filter(lead=lead, is_deleted=False).order_by('-timestamp')
        page = self.paginate_queryset(queryset)
        serializer = LeadHistorySerializer(page or queryset, many=True)
        if page is not None:
            return self.get_paginated_response(serializer.data)
        return Response(serializer.data)

    @extend_schema(
        summary="Create lead history entry",
        description="Manually create a history entry (optional)",
        tags=["Leads"],
    )
    @action(detail=True, methods=['post'])
    def add_history(self, request, pk=None):
        lead = self.get_object()
        data = request.data.copy()
        data['lead'] = lead.id
        serializer = LeadHistorySerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @extend_schema(
        summary="Delete lead history entry",
        tags=["Leads"],
    )
    @action(detail=False, methods=['delete'], url_path='history/(?P<history_id>[^/.]+)')
    def delete_history(self, request, history_id=None):
        try:
            entry = LeadHistory.objects.get(id=history_id)
            # Soft delete instead of hard delete
            if hasattr(entry, 'is_deleted'):
                LeadHistory.objects.filter(id=entry.id).update(is_deleted=True)
                return Response({"success": True, "message": "History entry deleted successfully"}, status=status.HTTP_200_OK)
            # Fallback if column not present
            entry.delete()
            return Response({"success": True, "message": "History entry deleted"}, status=status.HTTP_200_OK)
        except LeadHistory.DoesNotExist:
            return Response({"error": "History entry not found."}, status=status.HTTP_404_NOT_FOUND)

@extend_schema_view(
    list=extend_schema(tags=["Registration Groups"]),
    retrieve=extend_schema(tags=["Registration Groups"]),
    create=extend_schema(tags=["Registration Groups"]),
    update=extend_schema(tags=["Registration Groups"]),
    partial_update=extend_schema(tags=["Registration Groups"]),
    destroy=extend_schema(tags=["Registration Groups"]),
)
class RegistrationGroupViewSet(viewsets.ModelViewSet):
    queryset = RegistrationGroup.objects.filter(is_deleted=False)
    serializer_class = RegistrationGroupSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['name']
    search_fields = ['name']
    ordering_fields = ['name', 'created_at', 'updated_at']
    ordering = ['-created_at']

    def get_queryset(self):
        """
        Get queryset (excludes deleted records)
        """
        return RegistrationGroup.objects.filter(is_deleted=False)

    def get_serializer_class(self):
        return RegistrationGroupSerializer

    def list(self, request, *args, **kwargs):
        """
        List registration groups with optional filtering
        """
        queryset = self.filter_queryset(self.get_queryset())
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        """
        Retrieve a specific registration group
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        """
        Create a new registration group
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        registration_group = serializer.save()

        detail_serializer = RegistrationGroupSerializer(registration_group)
        return Response(detail_serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        """
        Update a specific registration group
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data)
        serializer.is_valid(raise_exception=True)
        registration_group = serializer.save()

        detail_serializer = RegistrationGroupSerializer(registration_group)
        return Response(detail_serializer.data, status=status.HTTP_200_OK)

    def destroy(self, request, *args, **kwargs):
        """
        Soft delete a registration group (sets is_deleted=True instead of actually deleting)
        """
        instance = self.get_object()
        instance.is_deleted = True
        instance.save()
        return Response(
            {"success": True, "message": "Registration group deleted successfully"},
            status=status.HTTP_200_OK
        )

@extend_schema_view(
    list=extend_schema(tags=["Lead Tags"]),
    retrieve=extend_schema(tags=["Lead Tags"]),
    create=extend_schema(tags=["Lead Tags"]),
    update=extend_schema(tags=["Lead Tags"]),
    partial_update=extend_schema(tags=["Lead Tags"]),
    destroy=extend_schema(tags=["Lead Tags"]),
)
class LeadTagViewSet(viewsets.ModelViewSet):
    queryset = LeadTag.objects.filter(is_deleted=False)
    serializer_class = LeadTagSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['name']
    search_fields = ['name']
    ordering_fields = ['name', 'created_at', 'updated_at']
    ordering = ['-created_at']

    def get_queryset(self):
        """
        Get queryset (excludes deleted records)
        """
        return LeadTag.objects.filter(is_deleted=False)

    def get_serializer_class(self):
        return LeadTagSerializer
    
    def list(self, request, *args, **kwargs):
        """
        List lead tags with optional filtering
        """
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        serializer = self.get_serializer(page or queryset, many=True)
        return self.get_paginated_response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        """
        Retrieve a specific lead tag
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        """
        Create a new lead tag
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        lead_tag = serializer.save()

        detail_serializer = LeadTagSerializer(lead_tag)
        return Response(detail_serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        """
        Update a specific lead tag
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data)
        serializer.is_valid(raise_exception=True)
        lead_tag = serializer.save()

        detail_serializer = LeadTagSerializer(lead_tag)
        return Response(detail_serializer.data, status=status.HTTP_200_OK)

    def destroy(self, request, *args, **kwargs):
        """
        Soft delete a lead tag (sets is_deleted=True instead of actually deleting)
        """
        instance = self.get_object()
        instance.is_deleted = True
        instance.save()
        return Response(
            {"success": True, "message": "Lead tag deleted successfully"},
            status=status.HTTP_200_OK
        )

@extend_schema_view(
    list=extend_schema(tags=["Sponsorship Types"]),
    retrieve=extend_schema(tags=["Sponsorship Types"]),
    create=extend_schema(tags=["Sponsorship Types"]),
    update=extend_schema(tags=["Sponsorship Types"]),
    partial_update=extend_schema(tags=["Sponsorship Types"]),
    destroy=extend_schema(tags=["Sponsorship Types"]),
)
class SponsorshipTypeViewSet(viewsets.ModelViewSet):
    queryset = SponsorshipType.objects.filter(is_deleted=False)
    serializer_class = SponsorshipTypeSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['name']
    search_fields = ['name']
    ordering_fields = ['name', 'created_at', 'updated_at']
    ordering = ['-created_at']

    def get_queryset(self):
        """
        Get queryset (excludes deleted records)
        """
        return SponsorshipType.objects.filter(is_deleted=False)

    def get_serializer_class(self):
        return SponsorshipTypeSerializer
    
    def list(self, request, *args, **kwargs):
        """
        List sponsorship types with optional filtering
        """
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        serializer = self.get_serializer(page or queryset, many=True)
        return self.get_paginated_response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        """
        Retrieve a specific sponsorship type
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        """
        Create a new sponsorship type
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        sponsorship_type = serializer.save()

        detail_serializer = SponsorshipTypeSerializer(sponsorship_type)
        return Response(detail_serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        """
        Update a specific sponsorship type
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data)
        serializer.is_valid(raise_exception=True)
        sponsorship_type = serializer.save()

        detail_serializer = SponsorshipTypeSerializer(sponsorship_type)
        return Response(detail_serializer.data, status=status.HTTP_200_OK)

    def destroy(self, request, *args, **kwargs):
        """
        Soft delete a sponsorship type (sets is_deleted=True instead of actually deleting)
        """
        instance = self.get_object()
        instance.is_deleted = True
        instance.save()
        return Response(
            {"success": True, "message": "Sponsorship type deleted successfully"},
            status=status.HTTP_200_OK
        )